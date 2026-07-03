# -*- coding: utf-8 -*-
"""
XTX NOR Flash Cross Reference 自动提取工具 V5（PDF逐列输出版）

V5核心逻辑：
1. 不再要求 PDF 型号必须匹配模板已有列；上传多少个 PDF，就输出多少个产品列。
2. 模板 A 列作为参数名、B 列作为填写示例时，最终自动删除 B 列。
3. 优先定位并解析 Command Table / DC Characteristics / AC Characteristics 三类表格。
4. 按模板参数行，把每份 PDF 的提取值写入对应型号列。
5. DC/AC 支持同一参数多温度合并到同一格，例如：
   ±2uA max@85C
   ±3uA max@105C
   ±4uA max@125C
6. 同时输出 Review_ACDC / Review_Command / PDF_Index，方便二次确认和优化规则。

运行：
    streamlit run cross_reference_streamlit_app_V5.py

依赖：
    streamlit==1.36.0
    openpyxl==3.1.5
    pymupdf==1.24.14
    pdfplumber==0.11.4
"""

from __future__ import annotations

import argparse
import copy
import io
import os
import re
import shutil
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import fitz  # PyMuPDF
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import pdfplumber
except Exception:
    pdfplumber = None

try:
    import streamlit as st
except Exception:
    st = None


VERSION = "V5.0 PDF逐列输出版"


# -----------------------------------------------------------------------------
# 基础规则：Excel规则文件会覆盖/补充这些默认规则
# -----------------------------------------------------------------------------
DEFAULT_ACDC_RULES: Dict[str, Dict[str, object]] = {
    # DC
    "Input Leakage Current": {"aliases": ["Input Leakage Current", "Input Leakage", "ILI", "I LI"], "units": ["uA", "µA", "μA"]},
    "Output Leakage Current": {"aliases": ["Output Leakage Current", "I/O Leakage", "IO Leakage", "Output Leakage", "ILO", "I LO"], "units": ["uA", "µA", "μA"]},
    "stanby_current": {"aliases": ["Standby Current", "Stand-by Current", "ICC1", "Icc1", "SB Current", "Standby"], "units": ["uA", "µA", "μA", "mA"]},
    "standby_current": {"aliases": ["Standby Current", "Stand-by Current", "ICC1", "Icc1", "SB Current", "Standby"], "units": ["uA", "µA", "μA", "mA"]},
    "deep_power_down_current": {"aliases": ["Deep Power-Down Current", "Deep Power Down Current", "Power-down Current", "Power Down Current", "ICC2", "Icc2", "DPD"], "units": ["uA", "µA", "μA"]},
    "read_current(Single)": {"aliases": ["Operating Current (Read)", "Current Read Data", "Read Current", "ICC3", "Icc3", "Serial Read", "Read 03H", "Single"], "units": ["mA", "uA", "µA", "μA"]},
    "read_current(Duad)": {"aliases": ["Dual Output", "Dual I/O", "Dual IO", "Dual Read", "ICC3", "Icc3"], "units": ["mA", "uA", "µA", "μA"]},
    "read_current(Dual)": {"aliases": ["Dual Output", "Dual I/O", "Dual IO", "Dual Read", "ICC3", "Icc3"], "units": ["mA", "uA", "µA", "μA"]},
    "read_current(Quad)": {"aliases": ["Quad Output", "Quad I/O", "Quad IO", "Quad Read", "ICC3", "Icc3"], "units": ["mA", "uA", "µA", "μA"]},
    "read_current(DTR Quad)": {"aliases": ["DTR Quad", "DDR Quad", "DTR", "DDR", "ICC3", "Icc3"], "units": ["mA", "uA", "µA", "μA"]},
    "program_current\nWRSR_current": {"aliases": ["Program Current", "Page Program", "Write Status Register", "WRSR", "ICC4", "ICC5", "Icc4", "Icc5"], "units": ["mA", "uA", "µA", "μA"]},
    "erase_current": {"aliases": ["Erase Current", "Sector Erase", "Block Erase", "Chip Erase", "ICC6", "ICC7", "ICC8", "Icc6", "Icc7", "Icc8"], "units": ["mA", "uA", "µA", "μA"]},
    "Input Low Voltage": {"aliases": ["Input Low Voltage", "VIL", "V IL"], "units": ["V"]},
    "Input High Voltage": {"aliases": ["Input High Voltage", "VIH", "V IH"], "units": ["V"]},
    "Output Low Voltage": {"aliases": ["Output Low Voltage", "VOL", "V OL"], "units": ["V"]},
    "Output High Voltage": {"aliases": ["Output High Voltage", "VOH", "V OH"], "units": ["V"]},

    # AC / timing
    "fC1(Serial Clock Frequency For: all commands except Read (03H))": {"aliases": ["fC1", "Serial Clock Frequency", "Clock Frequency", "SCLK Frequency"], "units": ["MHz"]},
    "fC2(Serial Clock Frequency For: DTR Read)": {"aliases": ["fC2", "DTR Read", "DDR Read", "DTR Clock", "DDR Clock"], "units": ["MHz"]},
    "fR(Serial Clock Frequency For: Read (03H))": {"aliases": ["fR", "Read Clock Frequency", "Read (03H)", "Read 03H"], "units": ["MHz"]},
    "tCLH(Serial Clock High Time)": {"aliases": ["tCLH", "tCH", "Serial Clock High Time", "Clock High Time"], "units": ["ns"]},
    "tCLL(Serial Clock Low Time)": {"aliases": ["tCLL", "tCL", "Serial Clock Low Time", "Clock Low Time"], "units": ["ns"]},
    "tCLCH(Serial Clock Rise/Fall Time (Slew Rate))": {"aliases": ["tCLCH", "Clock Rise Time", "Rise Time", "Slew Rate"], "units": ["ns", "V/ns", "Vns"]},
    "tCHCL(Serial Clock Rise/Fall Time (Slew Rate))": {"aliases": ["tCHCL", "Clock Fall Time", "Fall Time", "Slew Rate"], "units": ["ns", "V/ns", "Vns"]},
    "tSLCH(CS# Active Setup Time)": {"aliases": ["tSLCH", "tCSS", "CS# Active Setup Time", "CS Setup Time"], "units": ["ns"]},
    "tCHSH(CS# Active Hold Time)": {"aliases": ["tCHSH", "tCSH", "CS# Active Hold Time", "CS Hold Time"], "units": ["ns"]},
    "tCLSH(CS# Active Hold Time)": {"aliases": ["tCLSH", "CS# Active Hold Time", "CS Hold Time"], "units": ["ns"]},
    "tSHCH(CS# Not Active Setup Time)": {"aliases": ["tSHCH", "CS# Not Active Setup Time", "CS High Setup"], "units": ["ns"]},
    "tCHSL(CS# Not Active Hold Time)": {"aliases": ["tCHSL", "CS# Not Active Hold Time"], "units": ["ns"]},
    "tSHSL(CS# High Time)": {"aliases": ["tSHSL", "CS# High Time", "CS High Time"], "units": ["ns"]},
    "tSHQZ(Output Disable Time)": {"aliases": ["tSHQZ", "Output Disable Time", "Output Hi-Z"], "units": ["ns"]},
    "tCLQX(Output Hold Time)": {"aliases": ["tCLQX", "Output Hold Time"], "units": ["ns"]},
    "tCLQV(Clock Transient To Output Valid)": {"aliases": ["tCLQV", "Clock to Output Valid", "Output Valid"], "units": ["ns"]},
    "tDVCH(Data In Setup Time)": {"aliases": ["tDVCH", "Data In Setup Time", "Data Setup Time"], "units": ["ns"]},
    "tCHDX(Data In Hold Time)": {"aliases": ["tCHDX", "Data In Hold Time", "Data Hold Time"], "units": ["ns"]},

    # Program / erase timing
    "tW(write status register Cycle Time)": {"aliases": ["tW", "Write Status Register Cycle Time", "Write Status Register"], "units": ["ms", "us", "µs", "μs"]},
    "byte program(First)": {"aliases": ["Byte Program", "First Byte Program", "tBP"], "units": ["us", "µs", "μs", "ms"]},
    "byte program": {"aliases": ["Byte Program", "tBP"], "units": ["us", "µs", "μs", "ms"]},
    "page program(256byte)": {"aliases": ["Page Program", "256 byte", "256-byte", "tPP"], "units": ["ms", "us", "µs", "μs"]},
    "page program(512byte)": {"aliases": ["Page Program", "512 byte", "512-byte", "tPP"], "units": ["ms", "us", "µs", "μs"]},
    "4k_erase": {"aliases": ["4KB Sector Erase", "4 Kbyte", "4KB Erase", "Subsector Erase", "tSE"], "units": ["ms", "s"]},
    "32k_erase": {"aliases": ["32KB Block Erase", "32KB Erase", "tBE32"], "units": ["ms", "s"]},
    "64k_erase": {"aliases": ["64KB Block Erase", "64KB Erase", "Sector Erase", "tSE", "tBE"], "units": ["ms", "s"]},
    "256K_erase": {"aliases": ["256KB Block Erase", "256KB Erase", "tBE256"], "units": ["ms", "s"]},
    "512Mb bulk erase time": {"aliases": ["Bulk Erase", "Die Erase", "512Mb Bulk", "tBE"], "units": ["s"]},
    "chip_erase": {"aliases": ["Chip Erase", "Bulk Erase", "tCE"], "units": ["s"]},
}

DEFAULT_COMMAND_RULES: Dict[str, Dict[str, object]] = {
    "enabale_reset": {"aliases": ["Enable Reset", "Reset Enable", "RSTEN"]},
    "reset": {"aliases": ["Reset", "Software Reset", "Reset Memory", "RST"]},
    "Mode Bit Reset": {"aliases": ["Mode Bit Reset", "MBR"]},
    "NOP": {"aliases": ["No Operation", "NOP"]},
    "Legacy Software Reset": {"aliases": ["Legacy Software Reset"]},
    "read_manufacture_id": {"aliases": ["Read Manufacturer", "Read Manufacture", "Manufacturer ID", "Read ID"]},
    "read_jedec_id": {"aliases": ["Read JEDEC ID", "JEDEC ID", "Read Identification"]},
    "read_unique_id": {"aliases": ["Read Unique ID", "Unique ID"]},
    "read_sfpd": {"aliases": ["Read SFDP", "SFDP"]},
    "Read Electronic Signature": {"aliases": ["Read Electronic Signature", "RES"]},
    "READ STATUS REGISTER-1": {"aliases": ["Read Status Register-1", "Read Status Register", "RDSR", "Status Register-1"]},
    "READ STATUS REGISTER-2": {"aliases": ["Read Status Register-2", "Status Register-2"]},
    "READ STATUS REGISTER-3": {"aliases": ["Read Status Register-3", "Status Register-3"]},
    "WRITE ENABLE": {"aliases": ["Write Enable", "WREN"]},
    "WRITE DISABLE": {"aliases": ["Write Disable", "WRDI"]},
    "WRITE STATUS REGISTER-1": {"aliases": ["Write Status Register-1", "Write Status Register", "WRSR"]},
    "WRITE STATUS REGISTER-2": {"aliases": ["Write Status Register-2"]},
    "READ DATA": {"aliases": ["Read Data", "Read", "READ"]},
    "FAST READ": {"aliases": ["Fast Read"]},
    "DUAL OUTPUT FAST READ": {"aliases": ["Dual Output Fast Read", "Dual Output Read"]},
    "DUAL I/O FAST READ": {"aliases": ["Dual I/O Fast Read", "Dual IO Fast Read", "Dual I/O Read"]},
    "QUAD OUTPUT FAST READ": {"aliases": ["Quad Output Fast Read", "Quad Output Read"]},
    "QUAD I/O FAST READ": {"aliases": ["Quad I/O Fast Read", "Quad IO Fast Read", "Quad I/O Read"]},
    "DTR QUAD I/O READ": {"aliases": ["DTR Quad I/O Read", "DDR Quad I/O Read", "DTR Quad Read"]},
    "PAGE PROGRAM": {"aliases": ["Page Program"]},
    "QUAD PAGE PROGRAM": {"aliases": ["Quad Page Program", "Quad Input Page Program"]},
    "SECTOR ERASE 4KB": {"aliases": ["Sector Erase", "4KB Erase", "4K Erase"]},
    "BLOCK ERASE 32KB": {"aliases": ["Block Erase 32KB", "32KB Block Erase"]},
    "BLOCK ERASE 64KB": {"aliases": ["Block Erase 64KB", "64KB Block Erase"]},
    "CHIP ERASE": {"aliases": ["Chip Erase", "Bulk Erase"]},
}

SECTION_HINTS = ("DC Characteristics", "DC Electrical Characteristics", "AC Characteristics", "AC Electrical Characteristics", "Command", "Instruction")


@dataclass
class PdfInfo:
    pdf_path: Path
    vendor: str
    part: str
    file_name: str
    page_count: int
    dc_pages: str = ""
    ac_pages: str = ""
    cmd_pages: str = ""


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n")
    text = text.replace("\u00a0", " ").replace("µ", "u").replace("μ", "u").replace("℃", "C")
    text = text.replace("–", "-").replace("—", "-")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n+", "\n", text)
    return text.strip()


def one_line(value: object) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).strip()


def norm_key(value: object) -> str:
    s = one_line(value).lower()
    s = s.replace("duad", "dual")
    return re.sub(r"[^a-z0-9]+", "", s)


def split_multiline_cell(value: object) -> List[str]:
    s = clean_text(value)
    if not s:
        return []
    parts = re.split(r"[\n;]+", s)
    return [one_line(p) for p in parts if one_line(p)]


def is_enabled(value: object) -> bool:
    s = one_line(value).lower()
    return s not in {"", "0", "false", "n", "no", "否", "disable", "disabled"}


def load_excel_rules(rules_xlsx: Optional[Path]) -> Tuple[Dict[str, Dict[str, object]], Dict[str, Dict[str, object]]]:
    acdc = copy.deepcopy(DEFAULT_ACDC_RULES)
    cmd = copy.deepcopy(DEFAULT_COMMAND_RULES)
    if not rules_xlsx or not Path(rules_xlsx).exists():
        return acdc, cmd

    wb = load_workbook(rules_xlsx, data_only=True)

    if "ACDC_Rules" in wb.sheetnames:
        ws = wb["ACDC_Rules"]
        headers = {one_line(c.value): i + 1 for i, c in enumerate(ws[1]) if one_line(c.value)}
        for r in range(2, ws.max_row + 1):
            if not is_enabled(ws.cell(r, headers.get("Enable", 1)).value):
                continue
            item = one_line(ws.cell(r, headers.get("Item_Name", 2)).value)
            if not item:
                continue
            aliases = split_multiline_cell(ws.cell(r, headers.get("Aliases", 3)).value)
            units = split_multiline_cell(ws.cell(r, headers.get("Units", 4)).value)
            if aliases:
                acdc.setdefault(item, {"aliases": [], "units": []})
                old_aliases = list(acdc[item].get("aliases", []))
                acdc[item]["aliases"] = list(dict.fromkeys(old_aliases + aliases))
            if units:
                acdc.setdefault(item, {"aliases": [], "units": []})
                old_units = list(acdc[item].get("units", []))
                acdc[item]["units"] = list(dict.fromkeys(old_units + units))

    if "Command_Rules" in wb.sheetnames:
        ws = wb["Command_Rules"]
        headers = {one_line(c.value): i + 1 for i, c in enumerate(ws[1]) if one_line(c.value)}
        for r in range(2, ws.max_row + 1):
            if not is_enabled(ws.cell(r, headers.get("Enable", 1)).value):
                continue
            item = one_line(ws.cell(r, headers.get("Item_Name", 2)).value)
            if not item:
                continue
            aliases = split_multiline_cell(ws.cell(r, headers.get("Aliases", 3)).value)
            if aliases:
                cmd.setdefault(item, {"aliases": []})
                old_aliases = list(cmd[item].get("aliases", []))
                cmd[item]["aliases"] = list(dict.fromkeys(old_aliases + aliases))
    return acdc, cmd


def unpack_zip(zip_path: Path, work_dir: Path) -> Path:
    target = work_dir / "unzipped_specs"
    target.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "r") as zf:
        for member in zf.infolist():
            out = target / member.filename
            if not str(out.resolve()).startswith(str(target.resolve())):
                raise ValueError(f"压缩包中存在不安全路径：{member.filename}")
        zf.extractall(target)
    return target


def list_pdfs(spec_dir: Path) -> List[Path]:
    return sorted([p for p in spec_dir.rglob("*.pdf") if p.is_file()])


def vendor_from_path(pdf_path: Path, base_dir: Path) -> str:
    try:
        rel = pdf_path.relative_to(base_dir)
        if len(rel.parts) >= 2:
            return rel.parts[0]
    except Exception:
        pass
    name = pdf_path.name.upper()
    if name.startswith("GD"):
        return "GD"
    if name.startswith("W25") or "WINBOND" in name:
        return "Winbond"
    if name.startswith("MX") or "MACRONIX" in name or "MXIC" in name:
        return "MXIC"
    if name.startswith("MT") or "MICRON" in name:
        return "Micron"
    if name.startswith("S25") or name.startswith("S70") or "CYPRESS" in name or "SPANSION" in name:
        return "Spansion"
    if name.startswith("XM") or "XMC" in name:
        return "XMC"
    if name.startswith("EN") or "ESMT" in name:
        return "ESMT"
    return pdf_path.parent.name or "Unknown"


def normalize_part(token: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", token.upper())


def detect_part(pdf_path: Path, first_text: str = "") -> str:
    text = (pdf_path.stem + "\n" + first_text[:3000]).upper()
    patterns = [
        r"\bGD(?:25|55)[A-Z0-9\-_/\.]{5,32}\b",
        r"\bW25[A-Z0-9\-_/\.]{5,32}\b",
        r"\bMX(?:25|66)[A-Z0-9\-_/\.]{5,32}\b",
        r"\bMT25[A-Z0-9\-_/\.]{5,32}\b",
        r"\bS(?:25|70)[A-Z0-9\-_/\.]{5,32}\b",
        r"\bXM25[A-Z0-9\-_/\.]{5,32}\b",
        r"\bEN25[A-Z0-9\-_/\.]{5,32}\b",
    ]
    candidates: List[str] = []
    for pat in patterns:
        for m in re.finditer(pat, text):
            token = m.group(0).strip("_-./,;:()[]{}")
            if len(normalize_part(token)) >= 7:
                candidates.append(token)
    if candidates:
        # 优先选文件名中的候选；否则选最长
        stem_norm = normalize_part(pdf_path.stem)
        candidates = list(dict.fromkeys(candidates))
        candidates.sort(key=lambda x: (normalize_part(x) not in stem_norm, -len(normalize_part(x))))
        return candidates[0]
    return pdf_path.stem


def pdf_page_texts(pdf_path: Path, max_pages: Optional[int]) -> List[str]:
    texts: List[str] = []
    doc = fitz.open(str(pdf_path))
    page_count = doc.page_count if max_pages is None else min(doc.page_count, max_pages)
    for i in range(page_count):
        try:
            texts.append(clean_text(doc.load_page(i).get_text("text") or ""))
        except Exception:
            texts.append("")
    return texts


def page_temperature(text: str, fallback_index: int) -> str:
    s = one_line(text)
    # 优先提取 -40C~85C / -40°C to +105°C 等范围的上限温度
    patterns = [
        r"-\s*40\s*[°]?C\s*(?:~|to|-|～)\s*\+?\s*(85|105|125)\s*[°]?C",
        r"T[Aa]?\s*=\s*-\s*40\s*[°]?C\s*(?:~|to|-|～)\s*\+?\s*(85|105|125)\s*[°]?C",
        r"(85|105|125)\s*[°]?C",
    ]
    for pat in patterns:
        m = re.search(pat, s, re.IGNORECASE)
        if m:
            return f"{m.group(1)}C"
    defaults = ["85C", "105C", "125C"]
    return defaults[min(fallback_index, len(defaults) - 1)]


def classify_page(text: str) -> str:
    k = norm_key(text[:2500])
    if ("dccharacteristics" in k or "dcelectricalcharacteristics" in k or "dcelectricalspec" in k) and "symbol" in k:
        return "DC"
    if ("accharacteristics" in k or "acelectricalcharacteristics" in k or "actiming" in k) and ("symbol" in k or "parameter" in k):
        return "AC"
    if ("command" in k or "instruction" in k) and ("opcode" in k or "code" in k or "description" in k):
        return "CMD"
    return ""


def extract_tables(pdf_path: Path, max_pages: Optional[int], enable_pdfplumber: bool = True) -> List[Dict[str, object]]:
    """返回表格列表：{page, kind, temp, rows, raw_text}"""
    texts = pdf_page_texts(pdf_path, max_pages=max_pages)
    results: List[Dict[str, object]] = []
    section_count = {"DC": 0, "AC": 0, "CMD": 0}

    if enable_pdfplumber and pdfplumber is not None:
        try:
            with pdfplumber.open(str(pdf_path)) as pdf:
                n = len(pdf.pages) if max_pages is None else min(len(pdf.pages), max_pages)
                for i in range(n):
                    page_text = texts[i] if i < len(texts) else ""
                    page_kind = classify_page(page_text)
                    try:
                        tables = pdf.pages[i].extract_tables() or []
                    except Exception:
                        tables = []
                    for table in tables:
                        rows = []
                        for row in table or []:
                            cleaned = [one_line(c) for c in (row or [])]
                            if any(cleaned):
                                rows.append(cleaned)
                        if not rows:
                            continue
                        blob = " ".join(" | ".join(r) for r in rows[:5])
                        kind = page_kind or classify_page(blob)
                        if not kind:
                            # 某些 PDF 标题不在同页文字里，用表头判断
                            nk = norm_key(blob)
                            if "min" in nk and "typ" in nk and "max" in nk and "symbol" in nk:
                                kind = "DC" if any(x in norm_key(page_text) for x in ["dc", "current", "voltage"]) else "AC"
                            elif "opcode" in nk or "instruction" in nk or "command" in nk:
                                kind = "CMD"
                        if kind in {"DC", "AC", "CMD"}:
                            idx = section_count[kind]
                            section_count[kind] += 1
                            results.append({
                                "page": i + 1,
                                "kind": kind,
                                "temp": page_temperature(page_text, idx) if kind in {"DC", "AC"} else "",
                                "rows": rows,
                                "raw_text": page_text,
                            })
        except Exception:
            pass

    # 如果 pdfplumber 没提取到，使用文本行兜底
    if not results:
        for idx, text in enumerate(texts, start=1):
            kind = classify_page(text)
            if not kind:
                continue
            rows = [[one_line(x)] for x in text.splitlines() if one_line(x)]
            kidx = section_count[kind]
            section_count[kind] += 1
            results.append({
                "page": idx,
                "kind": kind,
                "temp": page_temperature(text, kidx) if kind in {"DC", "AC"} else "",
                "rows": rows,
                "raw_text": text,
            })
    return results


def find_header_map(rows: List[List[str]], kind: str) -> Tuple[Optional[int], Dict[str, int]]:
    """寻找表头并映射列：parameter/symbol/condition/min/typ/max/unit/opcode。"""
    best_idx = None
    best_score = 0
    best_map: Dict[str, int] = {}
    for i, row in enumerate(rows[:12]):
        mp: Dict[str, int] = {}
        score = 0
        for j, cell in enumerate(row):
            k = norm_key(cell)
            if not k:
                continue
            if k in {"parameter", "parameters"} or "parameter" in k:
                mp["parameter"] = j; score += 2
            if k in {"symbol", "symbols"} or "symbol" in k:
                mp["symbol"] = j; score += 2
            if "condition" in k or "testcondition" in k:
                mp["condition"] = j; score += 1
            if k in {"min", "minimum"} or k.endswith("min"):
                mp["min"] = j; score += 1
            if k in {"typ", "typical"} or k.endswith("typ"):
                mp["typ"] = j; score += 1
            if k in {"max", "maximum"} or k.endswith("max"):
                mp["max"] = j; score += 1
            if k in {"unit", "units"} or "unit" in k:
                mp["unit"] = j; score += 1
            if "opcode" in k or k in {"code", "instructioncode"}:
                mp["opcode"] = j; score += 2
            if "command" in k or "instruction" in k or "description" in k or "operation" in k:
                mp.setdefault("command", j); score += 1
        if kind in {"DC", "AC"}:
            # 支持 Winbond 的 SPEC 合并表头：第一行有 SPEC，第二行 MIN/TYP/MAX
            if {"min", "typ", "max"}.intersection(mp) and ("parameter" in mp or "symbol" in mp):
                score += 3
        if kind == "CMD" and ("opcode" in mp and "command" in mp):
            score += 3
        if score > best_score:
            best_score = score
            best_idx = i
            best_map = mp
    if best_score <= 1:
        return None, {}
    return best_idx, best_map


def get_cell(row: Sequence[str], idx: Optional[int]) -> str:
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    return one_line(row[idx])


def infer_unit(row: Sequence[str], header_map: Dict[str, int], blob: str = "") -> str:
    unit = get_cell(row, header_map.get("unit"))
    if unit:
        return unit.replace("μ", "u").replace("µ", "u")
    m = re.search(r"\b(uA|mA|A|V|MHz|kHz|Hz|ns|us|ms|s|pF)\b", blob.replace("µ", "u").replace("μ", "u"), re.I)
    return m.group(1) if m else ""


def normalize_number_value(v: str) -> str:
    v = one_line(v)
    if not v or v in {"-", "--", "—"}:
        return ""
    v = v.replace("μ", "u").replace("µ", "u")
    v = re.sub(r"\s+", "", v)
    # 保留 ±、<、>、VCC表达式
    return v


def extract_freq(condition: str) -> str:
    c = one_line(condition)
    # 166MHz、at 50MHz、104 MHz
    m = re.search(r"(\d+(?:\.\d+)?)\s*(MHz|KHz|kHz)", c, re.I)
    if m:
        return f"{m.group(1)}{m.group(2).replace('mhz','MHz').replace('MHZ','MHz')}"
    return ""


def format_min_typ_max(min_v: str, typ_v: str, max_v: str, unit: str, condition: str, temp: str) -> str:
    min_v = normalize_number_value(min_v)
    typ_v = normalize_number_value(typ_v)
    max_v = normalize_number_value(max_v)
    unit = normalize_number_value(unit)
    if unit.lower() in {"ua", "μa", "µa"}:
        unit = "uA"
    elif unit.lower() == "ma":
        unit = "mA"
    elif unit.lower() == "mhz":
        unit = "MHz"
    elif unit.lower() == "pf":
        unit = "pF"
    elif unit.lower() in {"us", "μs", "µs"}:
        unit = "us"

    freq = extract_freq(condition)
    suffix = ""
    if freq:
        suffix += f"@{freq}"
    if temp:
        suffix += f"@{temp}"

    # 只有 max，比如 ±2 + uA => ±2uA max@85C
    if max_v and not typ_v and not min_v:
        return f"{max_v}{unit} max{suffix}" if unit else f"{max_v} max{suffix}"
    if typ_v and max_v and not min_v:
        return f"{typ_v}{unit}/{max_v}{unit}{suffix}" if unit else f"{typ_v}/{max_v}{suffix}"
    if min_v and max_v and not typ_v:
        u = unit or ""
        return f"{min_v}{u} min/{max_v}{u} max{suffix}"
    if min_v and typ_v and max_v:
        u = unit or ""
        return f"{min_v}{u}/{typ_v}{u}/{max_v}{u}{suffix}"
    if typ_v and not max_v and not min_v:
        return f"{typ_v}{unit} typ{suffix}" if unit else f"{typ_v} typ{suffix}"
    if min_v and not typ_v and not max_v:
        return f"{min_v}{unit} min{suffix}" if unit else f"{min_v} min{suffix}"
    return ""


def is_section_row(value: str) -> bool:
    s = one_line(value)
    if not s:
        return True
    k = norm_key(s)
    if k in {"vendor", "partnumber", "item", "parameter", "templatevalue"}:
        return True
    if "characteristics" in k and len(s) < 80:
        return True
    if s.startswith("一、") or s.startswith("二、") or s.startswith("三、"):
        return True
    return False


def base_symbol_from_item(item: str) -> str:
    # 从 tCLH(...) / fC1(...) / VOL 这类模板项提取符号
    s = one_line(item)
    m = re.match(r"^([A-Za-z][A-Za-z0-9#_/-]{1,12})\s*\(", s)
    if m:
        return m.group(1)
    m = re.match(r"^(t[A-Za-z0-9]+|f[A-Za-z0-9]+|V[A-Z]{1,3}|I[A-Z]{1,3}\d?)\b", s)
    if m:
        return m.group(1)
    return ""


def row_matches_item(item: str, aliases: Sequence[str], symbol: str, parameter: str, condition: str) -> bool:
    blob = f"{symbol} {parameter} {condition}"
    nk = norm_key(blob)
    item_key = norm_key(item)
    # AC 模板常以 Symbol 开头，优先按 Symbol 精确匹配
    bs = base_symbol_from_item(item)
    if bs and norm_key(bs) and norm_key(bs) in norm_key(symbol):
        return True
    # DC特殊分流，避免 ICC3 所有行写进 Single
    lower_blob = one_line(blob).lower()
    lower_item = one_line(item).lower()
    if "quad" in lower_item and "dtr" not in lower_item and "quad" not in lower_blob:
        return False
    if "dual" in lower_item or "duad" in lower_item:
        if "dual" not in lower_blob:
            return False
    if "dtr" in lower_item and not ("dtr" in lower_blob or "ddr" in lower_blob):
        return False
    if "single" in lower_item:
        if "quad" in lower_blob or "dual" in lower_blob or "dtr" in lower_blob or "ddr" in lower_blob:
            return False
    for alias in aliases:
        ak = norm_key(alias)
        if ak and ak in nk:
            return True
    # 模板 Item 本身包含完整名称时也尝试匹配
    if item_key and item_key in nk:
        return True
    return False


def extract_acdc_from_tables(tables: List[Dict[str, object]], template_items: List[str], rules: Dict[str, Dict[str, object]]) -> Tuple[Dict[str, str], List[Dict[str, str]]]:
    found: Dict[str, List[str]] = {item: [] for item in template_items}
    evidences: List[Dict[str, str]] = []

    for table in tables:
        kind = str(table.get("kind", ""))
        if kind not in {"DC", "AC"}:
            continue
        rows: List[List[str]] = table.get("rows", [])  # type: ignore
        header_idx, hmap = find_header_map(rows, kind)
        if header_idx is None:
            continue

        # 对于 Winbond 的两层表头，如果当前行没有 parameter/symbol，尝试前一行+后一行综合。
        data_rows = rows[header_idx + 1:]
        temp = str(table.get("temp", ""))
        page = str(table.get("page", ""))
        last_param = ""
        last_symbol = ""

        for row in data_rows:
            blob = " | ".join([one_line(x) for x in row])
            if not one_line(blob):
                continue
            # 跳过重复表头/注释行
            if norm_key(blob) in {"minmaxunit", "spectypmaxunit", "parametertestconditionmintypmaxunit"}:
                continue

            parameter = get_cell(row, hmap.get("parameter"))
            symbol = get_cell(row, hmap.get("symbol"))
            condition = get_cell(row, hmap.get("condition"))
            min_v = get_cell(row, hmap.get("min"))
            typ_v = get_cell(row, hmap.get("typ"))
            max_v = get_cell(row, hmap.get("max"))
            unit = infer_unit(row, hmap, blob)

            # 合并单元格在 PDF 表格里可能为空，继承上一行 Parameter/Symbol
            if parameter:
                last_param = parameter
            else:
                parameter = last_param
            if symbol:
                last_symbol = symbol
            else:
                symbol = last_symbol

            value = format_min_typ_max(min_v, typ_v, max_v, unit, condition, temp)
            if not value:
                continue

            for item in template_items:
                if is_section_row(item):
                    continue
                rule = rules.get(item, {})
                aliases = list(rule.get("aliases", []))
                if row_matches_item(item, aliases, symbol, parameter, condition):
                    if value not in found[item]:
                        found[item].append(value)
                        evidences.append({
                            "Item": item,
                            "Value": value,
                            "Page": page,
                            "Kind": kind,
                            "Symbol": symbol,
                            "Parameter": parameter,
                            "Condition": condition,
                            "Raw_Row": blob,
                        })
    return {k: "\n".join(v) for k, v in found.items() if v}, evidences


def extract_opcode(text: str) -> str:
    s = one_line(text).upper()
    # 优先提取 9FH / 0x9F / 06h 这类命令码
    codes = []
    for pat in [r"\b0X([0-9A-F]{2})\b", r"\b([0-9A-F]{2})H\b", r"\b([0-9A-F]{2})\s*h\b"]:
        for m in re.finditer(pat, s, re.I):
            code = m.group(1).upper() + "H"
            if code not in codes:
                codes.append(code)
    # 表格中有时候直接写 06 / 9F
    if not codes:
        for m in re.finditer(r"(?:^|\s|\|)([0-9A-F]{2})(?:\s|\||$)", s, re.I):
            code = m.group(1).upper() + "H"
            if code not in codes:
                codes.append(code)
    return "/".join(codes[:3])


def extract_command_from_tables(tables: List[Dict[str, object]], template_items: List[str], rules: Dict[str, Dict[str, object]]) -> Tuple[Dict[str, str], List[Dict[str, str]]]:
    found: Dict[str, str] = {}
    evidences: List[Dict[str, str]] = []
    for table in tables:
        if table.get("kind") != "CMD":
            continue
        rows: List[List[str]] = table.get("rows", [])  # type: ignore
        page = str(table.get("page", ""))
        header_idx, hmap = find_header_map(rows, "CMD")
        start = (header_idx + 1) if header_idx is not None else 0
        for row in rows[start:]:
            blob = " | ".join([one_line(x) for x in row])
            if not one_line(blob):
                continue
            code = ""
            if hmap:
                code = get_cell(row, hmap.get("opcode"))
            if not code:
                code = extract_opcode(blob)
            else:
                code = extract_opcode(code) or code
            if not code:
                continue
            for item in template_items:
                if is_section_row(item):
                    continue
                aliases = list(rules.get(item, {}).get("aliases", []))
                nk = norm_key(blob)
                if any(norm_key(a) and norm_key(a) in nk for a in aliases) or norm_key(item) in nk:
                    found.setdefault(item, code)
                    evidences.append({"Item": item, "Opcode": code, "Page": page, "Raw_Row": blob})
    return found, evidences


def sheet_kind(ws) -> str:
    name = ws.title.lower()
    if "command" in name or "cmd" in name:
        return "CMD"
    if "acdc" in name or "ac" in name or "dc" in name or "character" in name:
        return "ACDC"
    # 从前30行判断
    blob = " ".join(one_line(ws.cell(r, c).value) for r in range(1, min(ws.max_row, 30) + 1) for c in range(1, min(ws.max_column, 6) + 1))
    if "Command" in blob or "Opcode" in blob:
        return "CMD"
    return "ACDC"


def score_item_col(ws, col: int) -> int:
    score = 0
    for r in range(1, min(ws.max_row, 120) + 1):
        v = one_line(ws.cell(r, col).value)
        if not v:
            continue
        k = norm_key(v)
        if any(x in k for x in ["inputleakage", "outputleakage", "standby", "deeppower", "readcurrent", "serialclock", "pagprogram", "chiperase", "writeenable", "readjedec", "fastread", "quad"]):
            score += 3
        if "characteristics" in k or "operation" in k:
            score += 1
    return score


def identify_template_layout(ws) -> Dict[str, int | bool]:
    # 判断 A列参数/B列示例 或 B列参数的老模板
    score_a = score_item_col(ws, 1)
    score_b = score_item_col(ws, 2)
    if score_a >= score_b:
        item_col = 1
        sample_col = 2
        product_start = 3
        delete_sample_b = True
    else:
        item_col = 2
        sample_col = 3 if ws.max_column >= 3 else 2
        product_start = 3
        delete_sample_b = False

    vendor_row = 1
    part_row = 2
    for r in range(1, min(ws.max_row, 10) + 1):
        v = norm_key(ws.cell(r, item_col).value)
        if "vendor" in v:
            vendor_row = r
        if "partnumber" in v or "partno" in v:
            part_row = r
    return {
        "item_col": item_col,
        "sample_col": sample_col,
        "product_start": product_start,
        "vendor_row": vendor_row,
        "part_row": part_row,
        "delete_sample_b": delete_sample_b,
    }


def collect_template_items(ws, item_col: int) -> Dict[int, str]:
    items: Dict[int, str] = {}
    for r in range(1, ws.max_row + 1):
        v = one_line(ws.cell(r, item_col).value)
        if v:
            items[r] = v
    return items


def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy.copy(src.protection)
    if src.hyperlink:
        dst._hyperlink = copy.copy(src.hyperlink)
    if src.comment:
        dst.comment = copy.copy(src.comment)


def prepare_product_columns(ws, layout: Dict[str, int | bool], pdf_count: int) -> None:
    product_start = int(layout["product_start"])
    sample_col = int(layout["sample_col"])
    # 删除模板原有产品列，避免旧数据干扰
    if ws.max_column >= product_start:
        ws.delete_cols(product_start, ws.max_column - product_start + 1)
    # 插入足够产品列
    if pdf_count > 0:
        ws.insert_cols(product_start, pdf_count)
    # 复制样式/宽度
    base_width = ws.column_dimensions[get_column_letter(sample_col)].width or 22
    for c in range(product_start, product_start + pdf_count):
        ws.column_dimensions[get_column_letter(c)].width = base_width
        for r in range(1, ws.max_row + 1):
            copy_cell_style(ws.cell(r, sample_col), ws.cell(r, c))
            ws.cell(r, c).alignment = copy.copy(ws.cell(r, sample_col).alignment)
            ws.cell(r, c).alignment = Alignment(
                horizontal=ws.cell(r, c).alignment.horizontal or "center",
                vertical=ws.cell(r, c).alignment.vertical or "center",
                wrap_text=True,
            )


def detect_pdf_info(pdf_path: Path, base_dir: Path, max_pages: Optional[int]) -> PdfInfo:
    try:
        doc = fitz.open(str(pdf_path))
        page_count = doc.page_count
        first_text = clean_text(doc.load_page(0).get_text("text") if page_count else "")
    except Exception:
        page_count = 0
        first_text = ""
    vendor = vendor_from_path(pdf_path, base_dir)
    part = detect_part(pdf_path, first_text)
    return PdfInfo(pdf_path=pdf_path, vendor=vendor, part=part, file_name=pdf_path.name, page_count=page_count)


def process_cross_reference(
    template_xlsx: Path,
    specs_zip: Path,
    output_xlsx: Path,
    rules_xlsx: Optional[Path] = None,
    max_pages: Optional[int] = 80,
    enable_pdfplumber: bool = True,
    delete_sample_b: bool = True,
    export_images_dir: Optional[Path] = None,
    progress=None,
) -> Dict[str, object]:
    work_dir = Path(tempfile.mkdtemp(prefix="crossref_v5_"))
    try:
        spec_dir = unpack_zip(specs_zip, work_dir)
        pdfs = list_pdfs(spec_dir)
        if not pdfs:
            raise ValueError("规格书 ZIP 中没有找到 PDF 文件。")

        acdc_rules, cmd_rules = load_excel_rules(rules_xlsx)
        wb = load_workbook(template_xlsx)

        pdf_infos: List[PdfInfo] = []
        pdf_tables: Dict[str, List[Dict[str, object]]] = {}
        pdf_acdc: Dict[str, Dict[str, str]] = {}
        pdf_cmd: Dict[str, Dict[str, str]] = {}
        acdc_reviews: List[Dict[str, str]] = []
        cmd_reviews: List[Dict[str, str]] = []

        # 先读取所有 PDF 表格
        for i, pdf in enumerate(pdfs, start=1):
            if progress:
                progress(f"读取 PDF：{i}/{len(pdfs)} - {pdf.name}")
            info = detect_pdf_info(pdf, spec_dir, max_pages)
            tables = extract_tables(pdf, max_pages=max_pages, enable_pdfplumber=enable_pdfplumber)
            info.dc_pages = ", ".join(str(t["page"]) for t in tables if t.get("kind") == "DC")
            info.ac_pages = ", ".join(str(t["page"]) for t in tables if t.get("kind") == "AC")
            info.cmd_pages = ", ".join(str(t["page"]) for t in tables if t.get("kind") == "CMD")
            pdf_infos.append(info)
            pdf_tables[str(pdf)] = tables

        written_cells = 0
        output_columns = len(pdf_infos)

        for ws in wb.worksheets:
            kind = sheet_kind(ws)
            layout = identify_template_layout(ws)
            item_col = int(layout["item_col"])
            sample_col = int(layout["sample_col"])
            product_start = int(layout["product_start"])
            vendor_row = int(layout["vendor_row"])
            part_row = int(layout["part_row"])
            should_delete_b = bool(layout["delete_sample_b"]) and delete_sample_b

            row_items = collect_template_items(ws, item_col)
            template_items = list(dict.fromkeys(v for v in row_items.values() if not is_section_row(v)))
            prepare_product_columns(ws, layout, output_columns)

            # 写表头
            for col_offset, info in enumerate(pdf_infos):
                col = product_start + col_offset
                ws.cell(vendor_row, col).value = info.vendor
                ws.cell(part_row, col).value = info.part or Path(info.file_name).stem
                ws.cell(vendor_row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.cell(part_row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # 对每个 PDF 针对本 Sheet 的模板项提取并写入
            for col_offset, info in enumerate(pdf_infos):
                col = product_start + col_offset
                tables = pdf_tables.get(str(info.pdf_path), [])
                if kind == "CMD":
                    extracted, evids = extract_command_from_tables(tables, template_items, cmd_rules)
                    pdf_cmd[str(info.pdf_path)] = extracted
                else:
                    extracted, evids = extract_acdc_from_tables(tables, template_items, acdc_rules)
                    pdf_acdc[str(info.pdf_path)] = extracted

                # 写入值
                for r, item in row_items.items():
                    if is_section_row(item):
                        continue
                    value = extracted.get(item, "")
                    if value:
                        ws.cell(r, col).value = value
                        ws.cell(r, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        written_cells += 1

                    template_value = one_line(ws.cell(r, sample_col).value)
                    review_row = {
                        "Sheet": ws.title,
                        "Vendor": info.vendor,
                        "Part_Number": info.part,
                        "PDF_File": info.file_name,
                        "Item": item,
                        "Template_Value": template_value,
                        "Extracted_Value": value,
                        "Match_Status": "Extracted" if value else "Missing",
                    }
                    if kind == "CMD":
                        cmd_reviews.append(review_row)
                    else:
                        acdc_reviews.append(review_row)

                # 附加 Evidence 页码信息到 Review
                for ev in evids:
                    ev_row = {
                        "Sheet": ws.title,
                        "Vendor": info.vendor,
                        "Part_Number": info.part,
                        "PDF_File": info.file_name,
                        **ev,
                    }
                    if kind == "CMD":
                        cmd_reviews.append(ev_row)
                    else:
                        acdc_reviews.append(ev_row)

            # 最后删除 B 列示例列：只针对 A列为参数、B列为模板示例的结构
            if should_delete_b and item_col == 1 and ws.max_column >= 2:
                ws.delete_cols(2, 1)

        # 创建 Review Sheet
        for name in ["Review_ACDC", "Review_Command", "PDF_Index"]:
            if name in wb.sheetnames:
                del wb[name]

        def write_review(sheet_name: str, rows: List[Dict[str, str]]):
            ws = wb.create_sheet(sheet_name)
            headers = []
            for row in rows:
                for k in row.keys():
                    if k not in headers:
                        headers.append(k)
            if not headers:
                headers = ["Sheet", "Vendor", "Part_Number", "PDF_File", "Item", "Template_Value", "Extracted_Value", "Match_Status"]
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(1, c, h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for r, row in enumerate(rows, start=2):
                for c, h in enumerate(headers, start=1):
                    ws.cell(r, c).value = row.get(h, "")
                    ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
            for c in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(c)].width = min(max(len(headers[c-1]) + 4, 16), 50)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        write_review("Review_ACDC", acdc_reviews)
        write_review("Review_Command", cmd_reviews)

        # PDF_Index
        idx_ws = wb.create_sheet("PDF_Index")
        idx_headers = ["No", "Vendor", "Part_Number", "PDF_File", "Page_Count", "DC_Pages", "AC_Pages", "Command_Pages", "Full_Path"]
        for c, h in enumerate(idx_headers, start=1):
            cell = idx_ws.cell(1, c, h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r, info in enumerate(pdf_infos, start=2):
            values = [r-1, info.vendor, info.part, info.file_name, info.page_count, info.dc_pages, info.ac_pages, info.cmd_pages, str(info.pdf_path)]
            for c, v in enumerate(values, start=1):
                idx_ws.cell(r, c).value = v
                idx_ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
        for c in range(1, len(idx_headers) + 1):
            idx_ws.column_dimensions[get_column_letter(c)].width = 24
        idx_ws.freeze_panes = "A2"
        idx_ws.auto_filter.ref = idx_ws.dimensions

        wb.save(output_xlsx)
        return {
            "pdf_count": len(pdf_infos),
            "output_columns": output_columns,
            "written_cells": written_cells,
            "review_acdc_rows": len(acdc_reviews),
            "review_command_rows": len(cmd_reviews),
            "output": str(output_xlsx),
        }
    finally:
        shutil.rmtree(work_dir, ignore_errors=True)


# -----------------------------------------------------------------------------
# Streamlit UI
# -----------------------------------------------------------------------------

def run_streamlit_app():
    st.set_page_config(page_title="XTX Cross Reference V5", layout="wide")
    st.title("XTX NOR Flash Cross Reference 自动提取工具 V5（PDF逐列输出版）")
    st.caption("V5逻辑：上传多少个PDF，就输出多少个产品列；不再要求PDF必须匹配模板已有型号列。")

    with st.sidebar:
        st.header("输入文件")
        template_file = st.file_uploader("Cross Reference 模板 Excel（.xlsx）", type=["xlsx"])
        specs_file = st.file_uploader("友商规格书压缩包（.zip，内容 PDF）", type=["zip"])
        rules_file = st.file_uploader("规则文件 Excel（可选，crossref_rules.xlsx）", type=["xlsx"])

        st.divider()
        st.header("提取设置")
        scan_opt = st.selectbox("PDF 扫描范围", ["仅前 80 页", "全文扫描", "仅前 120 页"], index=0)
        enable_pdfplumber = st.checkbox("启用 pdfplumber 表格增强抽取（推荐）", value=True)
        delete_b = st.checkbox("如果模板B列是填写示例，最终输出删除B列", value=True)
        out_name = st.text_input("输出文件名", value="Cross_Reference_Extracted_V5.xlsx")

        st.info("请确认页面标题包含 V5，且左侧不再有“PDF与模板列匹配阈值”。")

    st.subheader("处理逻辑")
    st.write("1. 复制模板并保留格式；2. 每个PDF直接生成一个产品列；3. 定位并解析 DC/AC/Command 表格；4. 按模板参数行写入对应值；5. 生成 Review 与 PDF_Index。")

    if not template_file or not specs_file:
        st.warning("请先上传模板 Excel 和规格书 ZIP。")
        return

    max_pages: Optional[int]
    if scan_opt == "全文扫描":
        max_pages = None
    elif scan_opt == "仅前 120 页":
        max_pages = 120
    else:
        max_pages = 80

    if st.button("开始生成", type="primary"):
        tmp = Path(tempfile.mkdtemp(prefix="crossref_v5_ui_"))
        try:
            template_path = tmp / template_file.name
            specs_path = tmp / specs_file.name
            rules_path = tmp / (rules_file.name if rules_file else "") if rules_file else None
            output_path = tmp / out_name

            template_path.write_bytes(template_file.getbuffer())
            specs_path.write_bytes(specs_file.getbuffer())
            if rules_file and rules_path:
                rules_path.write_bytes(rules_file.getbuffer())

            progress_box = st.empty()
            def progress(msg: str):
                progress_box.info(msg)

            with st.spinner("正在解析 PDF 并生成 Cross Reference，请稍候..."):
                result = process_cross_reference(
                    template_xlsx=template_path,
                    specs_zip=specs_path,
                    output_xlsx=output_path,
                    rules_xlsx=rules_path,
                    max_pages=max_pages,
                    enable_pdfplumber=enable_pdfplumber,
                    delete_sample_b=delete_b,
                    progress=progress,
                )
            progress_box.success("生成完成")

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("PDF 数量", result["pdf_count"])
            c2.metric("输出产品列", result["output_columns"])
            c3.metric("写入单元格", result["written_cells"])
            c4.metric("Review 行数", int(result["review_acdc_rows"]) + int(result["review_command_rows"]))

            data = output_path.read_bytes()
            st.download_button(
                "下载结果 Excel",
                data=data,
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as exc:
            st.error(f"生成失败：{exc}")
            st.exception(exc)
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


# -----------------------------------------------------------------------------
# CLI
# -----------------------------------------------------------------------------

def run_cli():
    parser = argparse.ArgumentParser(description="XTX Cross Reference Extractor V5")
    parser.add_argument("--template", required=True, help="Cross Reference 模板 .xlsx")
    parser.add_argument("--specs", required=True, help="友商规格书 .zip")
    parser.add_argument("--out", required=True, help="输出 .xlsx")
    parser.add_argument("--rules", default="", help="规则文件 crossref_rules.xlsx，可选")
    parser.add_argument("--max-pages", type=int, default=80, help="每份PDF扫描页数；0表示全文")
    parser.add_argument("--no-pdfplumber", action="store_true", help="禁用pdfplumber表格增强抽取")
    parser.add_argument("--keep-sample-b", action="store_true", help="保留B列模板示例，不删除")
    args = parser.parse_args()

    result = process_cross_reference(
        template_xlsx=Path(args.template),
        specs_zip=Path(args.specs),
        output_xlsx=Path(args.out),
        rules_xlsx=Path(args.rules) if args.rules else None,
        max_pages=None if args.max_pages == 0 else args.max_pages,
        enable_pdfplumber=not args.no_pdfplumber,
        delete_sample_b=not args.keep_sample_b,
        progress=print,
    )
    print(result)


if __name__ == "__main__":
    if "--template" in os.sys.argv and "--specs" in os.sys.argv:
        run_cli()
    else:
        if st is None:
            raise RuntimeError("请先安装 streamlit：pip install streamlit")
        run_streamlit_app()
